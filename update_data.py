"""
Gabung data cabang lalu unggah ke GitHub — sekali jalan.
=======================================================

Alur kerjanya:
  1. Membaca semua berkas .xlsx di folder masukan
  2. Mengenali jenisnya sendiri (data servis atau data penjualan)
  3. Menggabungkan, membuang baris kembar, menyeragamkan nama cabang
  4. Menulis hasilnya ke folder data/ pada repo
  5. Commit & push ke GitHub

Pemakaian:
    python update_data.py                    # pakai pengaturan di config.json
    python update_data.py --tanpa-push       # gabung saja, tidak diunggah
    python update_data.py --masukan ~/Downloads/agustus

Pengaturan disimpan di config.json (dibuat otomatis saat pertama dijalankan).
"""
import argparse
import csv
import gzip
import json
import os
import re
import subprocess
import sys
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    print("Pustaka openpyxl belum terpasang. Jalankan:  pip install openpyxl")
    sys.exit(1)

BASE = Path(__file__).parent.resolve()
CONFIG = BASE / "config.json"

BAWAAN = {
    "folder_masukan": "masukan",
    "folder_repo": "",
    "subfolder_data": "data",
    "berkas_servis": "latest_data.csv.gz",
    "berkas_penjualan": "penjualan.csv.gz",
    "cabang_alias": {
        "TELUK JAMBE": "KARAWANG",
        "TELUKJ": "KARAWANG",
    },
    "branch": "main",
}

SHEET_SERVIS = "Rincian Pengiriman Pesanan"
SHEET_JUAL = "Rincian Faktur Penjualan"

# potongan nama pada berkas hasil export -> nama cabang
KODE_CABANG = {
    "klende": "KLENDER", "cegert": "CEGER", "ceger": "CEGER",
    "bintar": "BINTARA", "radjim": "RADJIMAN", "jatimu": "JATIMULYA",
    "dramag": "DRAMAGA", "condet": "CONDET", "jatibe": "JATIBENING",
    "sawang": "SAWANGAN", "warbon": "WARBONG", "cinere": "CINERE",
    "cibino": "CIBINONG", "telukj": "KARAWANG", "karawa": "KARAWANG",
    "jatiwa": "JATIWARINGIN", "cikamp": "CIKAMPEK", "cilang": "CILANGKAP",
    "pejate": "PEJATEN", "cibubu": "CIBUBUR",
}


# ---------------------------------------------------------------------------
# pengaturan
# ---------------------------------------------------------------------------
def muat_config():
    if CONFIG.exists():
        cfg = dict(BAWAAN)
        cfg.update(json.loads(CONFIG.read_text(encoding="utf-8")))
        return cfg
    CONFIG.write_text(json.dumps(BAWAAN, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Berkas pengaturan dibuat: {CONFIG}")
    print("Buka berkas itu, isi 'folder_repo' dengan lokasi repo GitHub Anda,")
    print("lalu jalankan ulang perintah ini.\n")
    sys.exit(0)


def nama_cabang(path, cfg):
    """Tentukan nama cabang dari nama berkas."""
    nama = Path(path).stem
    m = re.search(r"_(\d{3})mflash([a-z]+)_", nama)
    if m:
        kode = m.group(2)
        cab = KODE_CABANG.get(kode, kode.upper())
    else:
        cab = re.sub(r"-[0-9a-f]{6,}$", "", nama).upper()
    alias = {k.upper(): v.upper() for k, v in cfg.get("cabang_alias", {}).items()}
    return alias.get(cab, cab)


# ---------------------------------------------------------------------------
# pembacaan
# ---------------------------------------------------------------------------
def baca_berkas(path, sheet):
    """Baca satu berkas Excel; kembalikan (header, baris) tanpa kolom kosong."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None, None
    ws = wb[sheet]
    ws.reset_dimensions()          # metadata dimensi berkas ini kerap tidak akurat

    it = ws.iter_rows(values_only=True)
    raw = next(it)
    keep = [i for i, c in enumerate(raw) if c is not None and str(c).strip()]
    header = [str(raw[i]).strip() for i in keep]

    baris = []
    for row in it:
        if row is None:
            continue
        vals = [row[i] if i < len(row) else None for i in keep]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue
        baris.append(vals)
    wb.close()
    return header, baris


def kunci(vals, cabang):
    return (cabang,) + tuple(
        v.isoformat() if isinstance(v, datetime) else v for v in vals)


def gabungkan(berkas, sheet, buang_kembar, cfg):
    """Gabungkan banyak berkas jadi satu; kembalikan (header, baris, ringkasan)."""
    header_acuan = None
    semua = []
    ringkas = []
    seen = set()

    for path in berkas:
        cab = nama_cabang(path, cfg)
        header, baris = baca_berkas(path, sheet)
        if header is None:
            continue
        if header_acuan is None:
            header_acuan = header
        elif header != header_acuan:
            beda = [(a, b) for a, b in zip(header_acuan, header) if a != b]
            if beda:
                print(f"    catatan: nama kolom {cab} sedikit berbeda "
                      f"({beda[0][1]} -> dipakai {beda[0][0]})")

        n_awal = len(baris)
        dipakai = 0
        for vals in baris:
            if buang_kembar:
                k = kunci(vals, cab)
                if k in seen:
                    continue
                seen.add(k)
            semua.append(vals + [cab])
            dipakai += 1
        ringkas.append((cab, n_awal, dipakai))
        print(f"  {cab:14s} {n_awal:7,} baris -> {dipakai:7,}")

    return header_acuan, semua, ringkas


# ---------------------------------------------------------------------------
# penulisan
# ---------------------------------------------------------------------------
def tulis_gz(path, header, baris):
    def fmt(v):
        if v is None:
            return ""
        if isinstance(v, datetime):
            if v.hour or v.minute or v.second:
                return v.strftime("%Y-%m-%d %H:%M:%S")
            return v.strftime("%Y-%m-%d")
        return v

    path.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(path, "wt", newline="", encoding="utf-8", compresslevel=9) as f:
        w = csv.writer(f)
        w.writerow(header + ["CABANG"])
        w.writerows([fmt(v) for v in r] for r in baris)
    return path.stat().st_size


# ---------------------------------------------------------------------------
# git
# ---------------------------------------------------------------------------
def jalankan_git(repo, *args):
    hasil = subprocess.run(["git", "-C", str(repo)] + list(args),
                           capture_output=True, text=True)
    return hasil.returncode, (hasil.stdout + hasil.stderr).strip()


def push_ke_github(repo, berkas_relatif, pesan, branch):
    kode, keluar = jalankan_git(repo, "rev-parse", "--is-inside-work-tree")
    if kode != 0:
        print(f"\n! Folder repo tidak dikenali sebagai repo Git: {repo}")
        print("  Pastikan 'folder_repo' di config.json menunjuk hasil 'git clone'.")
        return False

    # pakai nama branch yang benar-benar aktif, bukan tebakan dari pengaturan
    kode, aktif = jalankan_git(repo, "branch", "--show-current")
    if kode == 0 and aktif.strip():
        if aktif.strip() != branch:
            print(f"  (branch aktif '{aktif.strip()}', bukan '{branch}' — memakai yang aktif)")
        branch = aktif.strip()

    kode, remote = jalankan_git(repo, "remote")
    if kode != 0 or not remote.strip():
        print("\n! Repo ini belum terhubung ke GitHub (tidak ada remote).")
        print("  Hubungkan sekali dengan:")
        print("    git remote add origin https://github.com/NAMA-ANDA/NAMA-REPO.git")
        return False

    for b in berkas_relatif:
        jalankan_git(repo, "add", b)

    kode, keluar = jalankan_git(repo, "diff", "--cached", "--stat")
    if not keluar.strip():
        print("\n= Tidak ada perubahan pada data. Tidak perlu push.")
        return True
    print(f"\nPerubahan yang akan diunggah:\n{keluar}")

    kode, keluar = jalankan_git(repo, "commit", "-m", pesan)
    if kode != 0 and "nothing to commit" not in keluar.lower():
        print(f"! Gagal commit:\n{keluar}")
        return False

    print("Mengunggah ke GitHub...")
    kode, keluar = jalankan_git(repo, "push", "origin", branch)
    if kode != 0:
        print(f"! Gagal push:\n{keluar}\n")
        print("Penyebab yang sering terjadi:")
        print("  - Belum login. Jalankan sekali:  git config --global credential.helper store")
        print("    lalu push manual sekali dan masukkan username + Personal Access Token.")
        print("  - Nama branch berbeda. Cek dengan:  git branch --show-current")
        return False

    print("Berhasil diunggah ke GitHub.")
    return True


# ---------------------------------------------------------------------------
# utama
# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(description="Gabung data cabang lalu unggah ke GitHub")
    p.add_argument("--masukan", help="folder berisi berkas .xlsx")
    p.add_argument("--repo", help="folder repo GitHub hasil git clone")
    p.add_argument("--tanpa-push", action="store_true", help="gabung saja, jangan unggah")
    p.add_argument("--pesan", help="pesan commit")
    a = p.parse_args()

    cfg = muat_config()
    masukan = Path(a.masukan or cfg["folder_masukan"])
    if not masukan.is_absolute():
        masukan = BASE / masukan
    repo = Path(a.repo or cfg["folder_repo"]).expanduser() if (a.repo or cfg["folder_repo"]) else None

    if not masukan.exists():
        masukan.mkdir(parents=True, exist_ok=True)
        print(f"Folder masukan dibuat: {masukan}")
        print("Letakkan berkas .xlsx dari sistem ke folder itu, lalu jalankan lagi.")
        return 0

    berkas = sorted(masukan.glob("*.xlsx"))
    berkas = [b for b in berkas if not b.name.startswith("~$")]
    if not berkas:
        print(f"Tidak ada berkas .xlsx di {masukan}")
        return 1

    print(f"Ditemukan {len(berkas)} berkas di {masukan}\n")

    # kelompokkan menurut jenis sheet
    servis, jual, tak_dikenal = [], [], []
    for b in berkas:
        try:
            wb = openpyxl.load_workbook(b, data_only=True, read_only=True)
            sh = wb.sheetnames
            wb.close()
        except Exception as e:
            print(f"  ! {b.name} tidak terbaca: {e}")
            tak_dikenal.append(b)
            continue
        if SHEET_SERVIS in sh:
            servis.append(b)
        elif SHEET_JUAL in sh:
            jual.append(b)
        else:
            tak_dikenal.append(b)

    if tak_dikenal:
        print("Berkas yang jenisnya tidak dikenali (dilewati):")
        for b in tak_dikenal:
            print(f"  - {b.name}")
        print()

    tujuan = (repo / cfg["subfolder_data"]) if repo else (BASE / "hasil")
    dibuat = []

    if servis:
        print(f"DATA SERVIS ({len(servis)} berkas) — baris kembar dibuang:")
        hdr, baris, ringkas = gabungkan(servis, SHEET_SERVIS, True, cfg)
        if baris:
            path = tujuan / cfg["berkas_servis"]
            ukuran = tulis_gz(path, hdr, baris)
            tot_awal = sum(r[1] for r in ringkas)
            print(f"  TOTAL {tot_awal:,} baris mentah -> {len(baris):,} transaksi unik "
                  f"({tot_awal - len(baris):,} kembar dibuang)")
            print(f"  Tersimpan: {path}  ({ukuran/1024/1024:.1f} MB)\n")
            dibuat.append(path)

    if jual:
        print(f"DATA PENJUALAN ({len(jual)} berkas) — semua baris dipakai:")
        hdr, baris, ringkas = gabungkan(jual, SHEET_JUAL, False, cfg)
        if baris:
            path = tujuan / cfg["berkas_penjualan"]
            ukuran = tulis_gz(path, hdr, baris)
            print(f"  TOTAL {len(baris):,} baris")
            print(f"  Tersimpan: {path}  ({ukuran/1024/1024:.1f} MB)\n")
            dibuat.append(path)

    if not dibuat:
        print("Tidak ada data yang berhasil digabung.")
        return 1

    if a.tanpa_push or not repo:
        if not repo:
            print("Catatan: 'folder_repo' belum diisi di config.json, jadi hasil "
                  f"disimpan di {tujuan} dan tidak diunggah.")
        return 0

    relatif = [str(p.relative_to(repo)) for p in dibuat]
    pesan = a.pesan or f"Perbarui data {datetime.now():%d %B %Y}"
    berhasil = push_ke_github(repo, relatif, pesan, cfg.get("branch", "main"))
    return 0 if berhasil else 1


if __name__ == "__main__":
    sys.exit(main())
